from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
import sqlite3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import secrets
import os
import traceback

app = Flask(__name__)
CORS(app)
app.secret_key = secrets.token_hex(16)
DATABASE = 'anwesenheit.db'

def init_db():
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS teilnahmen (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            datum TEXT NOT NULL,
            uhrzeit TEXT NOT NULL,
            kurs TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/teilnahme', methods=['POST'])
def teilnahme_bestaetigen():
    try:
        if not request.is_json:
            return jsonify({'message': 'Anfrage ist kein JSON!'}), 400

        data = request.get_json()
        name = data.get('name')
        kurs = data.get('kurs')

        if not name or not kurs:
            return jsonify({'message': 'Name oder Kurs fehlen'}), 400

        datum = datetime.now().strftime('%Y-%m-%d')
        uhrzeit = datetime.now().strftime('%H:%M:%S')

        conn = sqlite3.connect(DATABASE)
        c = conn.cursor()
        c.execute('INSERT INTO teilnahmen (name, datum, uhrzeit, kurs) VALUES (?, ?, ?, ?)',
                  (name, datum, uhrzeit, kurs))
        conn.commit()
        conn.close()

        return jsonify({'message': f'Teilnahme von {name} am {datum} um {uhrzeit} gespeichert (Kurs: {kurs}).'})

    except Exception as e:
        traceback.print_exc()
        return jsonify({'message': f'Fehler im Server: {str(e)}'}), 500

@app.route('/api/teilnehmerliste', methods=['GET'])
def get_teilnehmerliste():
    datum = datetime.now().strftime('%Y-%m-%d')
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute('SELECT name, kurs FROM teilnahmen WHERE datum = ?', (datum,))
    daten = c.fetchall()
    conn.close()
    return jsonify([{"name": name, "kurs": kurs} for name, kurs in daten])

@app.route('/download', methods=['GET'])
def download_liste():
    datum = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute('SELECT name, datum, uhrzeit, kurs FROM teilnahmen WHERE datum = ?', (datum,))
    daten = c.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Teilnehmer {datum}"

    header = ["Name", "Datum", "Uhrzeit", "Kurs"]
    ws.append(header)

    fett = Font(bold=True)
    zentriert = Alignment(horizontal="center")
    rahmen = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for zelle in ws[1]:
        zelle.font = fett
        zelle.alignment = zentriert
        zelle.border = rahmen

    for eintrag in daten:
        ws.append(eintrag)

    for zeile in ws.iter_rows(min_row=2):
        for zelle in zeile:
            zelle.alignment = zentriert
            zelle.border = rahmen

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    dateiname = f"anwesenheit_{datum}.xlsx"
    wb.save(dateiname)

    return send_file(
        dateiname,
        as_attachment=True,
        download_name=dateiname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host="0.0.0.0")
